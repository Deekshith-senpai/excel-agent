import openpyxl
from app.value_parser import parse_value
from app.llm_mapper import map_columns_with_llm
from app.models import ParsedCell


def validate_value(param_name, parsed_value, row_index):
    """
    Returns validation issue dict if invalid, else None
    """

    if parsed_value is None:
        return None

    # Rule 1: Coal consumption must be >= 0
    if param_name == "coal_consumption":
        if parsed_value < 0:
            return {
                "row": row_index,
                "parameter": param_name,
                "issue": "negative_value",
                "message": "Coal consumption cannot be negative"
            }

    # Rule 2: Steam generation must be >= 0
    if param_name == "steam_generation":
        if parsed_value < 0:
            return {
                "row": row_index,
                "parameter": param_name,
                "issue": "negative_value",
                "message": "Steam generation cannot be negative"
            }

    # Rule 3: Efficiency must be between 0 and 1
    if param_name == "efficiency":
        if not (0 <= parsed_value <= 1):
            return {
                "row": row_index,
                "parameter": param_name,
                "issue": "out_of_range",
                "message": "Efficiency must be between 0% and 100%"
            }

    return None

def detect_header_row(sheet):
    """
    Smart header detection:
    - Checks first 10 rows
    - Looks for at least 2 text-like cells
    - Skips numeric-only rows
    """
    for row_index in range(1, 11):  # 1-based indexing in openpyxl
        row = sheet[row_index]

        values = [
            str(cell.value).strip().lower()
            for cell in row
            if cell.value is not None
        ]

        if not values:
            continue

        text_cells = [
            v for v in values
            if not v.replace(".", "").replace(",", "").isdigit()
        ]

        if len(text_cells) >= 2:
            return row_index - 1  # convert to 0-based index

    return 0


async def parse_excel(file):

    file.file.seek(0)
    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active

    warnings = []
    validation_warnings = []
    duplicates = []

    # 🔥 Smart header detection
    header_row_index = detect_header_row(sheet)

    if header_row_index > 0:
        warnings.append(
            f"Rows 1-{header_row_index} appear to be title rows, skipped"
        )

    headers = [
        cell.value for cell in sheet[header_row_index + 1]
        if cell.value is not None
    ]

    columns = [str(h).strip() for h in headers]

    # 🔥 Map columns using LLM
    mapping = map_columns_with_llm(columns)

    parsed_data = []
    unmapped = []

    # 🔥 DUPLICATE DETECTION (COLUMN LEVEL)
    seen_param_asset = set()

    for map_result in mapping:

        column_name = map_result.get("column")

        if column_name not in columns:
            continue

        col_index = columns.index(column_name)

        param_name = map_result.get("param_name")
        asset_name = map_result.get("asset_name")

        # Unmapped column handling
        if not param_name:
            unmapped.append({
                "col": col_index,
                "header": column_name,
                "reason": "No matching parameter found"
            })
            continue

        # 🔥 Duplicate detection (same param+asset in multiple columns)
        key = (param_name, asset_name)

        if key in seen_param_asset:
            duplicates.append({
                "param_name": param_name,
                "asset_name": asset_name,
                "column": column_name,
                "reason": "Duplicate parameter+asset combination detected"
            })
        else:
            seen_param_asset.add(key)

        # 🔥 Parse rows
        for row_index in range(header_row_index + 2, sheet.max_row + 1):

            raw_value = sheet.cell(
                row=row_index,
                column=col_index + 1
            ).value

            if raw_value is None or str(raw_value).strip() == "":
                continue

            parsed_value = parse_value(raw_value)

            # -----------------------------
            # 🔥 VALIDATION RULES
            # -----------------------------
            issue = validate_value(param_name, parsed_value, row_index)

            if issue:
                validation_warnings.append(issue)

            parsed_data.append(
                ParsedCell(
                    row=row_index,
                    col=col_index,
                    param_name=param_name,
                    asset_name=asset_name,
                    raw_value=str(raw_value),
                    parsed_value=parsed_value,
                    confidence=map_result.get("confidence", "medium")
                )
            )

    return {
        "status": "success",
        "header_row": header_row_index,
        "parsed_data": parsed_data,
        "unmapped_columns": unmapped,
        "warnings": warnings,
        "validation_issues": validation_warnings,
        "duplicates": duplicates
    }