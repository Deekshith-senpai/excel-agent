from openpyxl import Workbook
import os

os.makedirs("test_data", exist_ok=True)

# ---------------------------------------------------
# 1️⃣ CLEAN DATA (Baseline - Should Parse Perfectly)
# ---------------------------------------------------
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Clean Data"

headers_clean = [
    "Coal Consumption",
    "Steam Generation",
    "Power Generation",
    "Efficiency"
]

ws1.append(headers_clean)

for i in range(10):
    ws1.append([
        1000 + i * 10,
        50 + i,
        20 + i,
        85 + i
    ])

wb1.save("test_data/clean_data.xlsx")


# ---------------------------------------------------
# 2️⃣ MESSY DATA (Header Detection + Fuzzy Matching)
# ---------------------------------------------------
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Messy Data"

ws2.append(["MONTHLY PLANT REPORT - JAN"])
ws2.append(["Generated Automatically"])
ws2.append([])

headers_messy = [
    "Coal Used",
    "Steam Gen (T/hr)",
    "Power (MWh)",
    "Boiler Eff %",
    "Comments"
]

ws2.append(headers_messy)

for i in range(10):
    ws2.append([
        f"{1_000 + i*20:,}",     # Comma format
        f"{50 + i}%",            # Percentage format
        f"{20 + i}",
        f"{85 + i}%",
        "OK" if i % 2 == 0 else "Check"
    ])

ws2.append([])
ws2.append(["End of Report"])

wb2.save("test_data/messy_data.xlsx")


# ---------------------------------------------------
# 3️⃣ MULTI ASSET DATA (Asset Detection + Deduplication)
# ---------------------------------------------------
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "Multi Asset"

headers_multi = [
    "Coal AFBC-1",
    "Coal AFBC-2",
    "Steam AFBC-1",
    "Steam AFBC-2",
    "Power TG-1",
    "Power TG-2",
    "Boiler Eff AFBC-1",
    "Boiler Eff AFBC-2"
]

ws3.append(headers_multi)

for i in range(10):
    ws3.append([
        1000 + i * 10,
        950 + i * 12,
        50 + i,
        48 + i,
        20 + i,
        18 + i,
        85 + i,
        83 + i
    ])

wb3.save("test_data/multi_asset.xlsx")

print("✅ All test files created inside test_data folder.")